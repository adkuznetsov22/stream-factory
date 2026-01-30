from __future__ import annotations

from io import BytesIO
from typing import Annotated

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import case, func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.youtube_sync import sync_youtube_account
from app.services.vk_sync import sync_vk_account
from app.services.tiktok_sync import sync_tiktok_account
from app.services.instagram_sync import sync_instagram_account
from .db import get_session
from .models import (
    AccountMetricsDaily,
    AccountOnboarding,
    Email,
    Phone,
    SocialAccount,
    SocialPlatform,
    VKPost,
    VKProfile,
    YouTubeChannel,
    TikTokProfile,
    InstagramProfile,
)
from .schemas import (
    EmailCreate,
    EmailRead,
    PhoneCreate,
    PhoneRead,
    SocialAccountCreate,
    SocialAccountRead,
    SocialAccountUpdate,
)

router = APIRouter(prefix="/api/accounts", tags=["accounts"])
phones_router = APIRouter(prefix="/api/phones", tags=["phones"])
emails_router = APIRouter(prefix="/api/emails", tags=["emails"])


SessionDep = Annotated[AsyncSession, Depends(get_session)]


def normalize_login(value: str) -> str:
    return value.strip().replace(" ", "").lstrip("@").lower()


def normalize_vk_slug(value: str) -> str:
    raw = (value or "").strip()
    raw = raw.replace("https://", "").replace("http://", "")
    raw = raw.replace("vk.com/", "").replace("m.vk.com/", "")
    raw = raw.strip("/").lstrip("@")
    return normalize_login(raw)


def build_url(platform: SocialPlatform, login: str, url: str | None) -> str:
    if platform == SocialPlatform.vk:
        slug = normalize_vk_slug(url or login)
        if not slug:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="VK url required")
        return f"https://vk.com/{slug}"
    if url:
        return url
    if platform == SocialPlatform.youtube:
        return f"https://www.youtube.com/@{login}"
    if platform == SocialPlatform.tiktok:
        return f"https://www.tiktok.com/@{login}"
    if platform == SocialPlatform.instagram:
        return f"https://www.instagram.com/{login}/"
    return url or ""


@router.get("", response_model=list[SocialAccountRead])
async def list_accounts(
    session: SessionDep,
    platform: SocialPlatform | None = Query(default=None),
    q: str | None = Query(default=None),
) -> list[SocialAccountRead]:
    metrics_agg = (
        select(
            AccountMetricsDaily.account_id.label("m_account_id"),
            func.sum(
                func.coalesce(
                    case(
                        (AccountMetricsDaily.date >= func.current_date() - 1, AccountMetricsDaily.views),
                        else_=0,
                    ),
                )
            ).label("views_24h"),
            func.sum(
                func.coalesce(
                    case(
                        (AccountMetricsDaily.date >= func.current_date() - 7, AccountMetricsDaily.views),
                        else_=0,
                    ),
                )
            ).label("views_7d"),
            func.sum(func.coalesce(AccountMetricsDaily.posts, 0)).label("posts_total"),
            func.max(AccountMetricsDaily.date).label("last_post_at"),
        )
        .group_by(AccountMetricsDaily.account_id)
        .subquery()
    )

    vk_posts_max = (
        select(VKPost.account_id.label("vk_account_id"), func.max(VKPost.published_at).label("vk_last_post_at"))
        .group_by(VKPost.account_id)
        .subquery()
    )
    yt_avatar = (
        select(YouTubeChannel.account_id.label("yt_account_id"), YouTubeChannel.thumbnail_url.label("yt_thumb"))
        .subquery()
    )
    vk_avatar = (
        select(VKProfile.account_id.label("vk_profile_account_id"), VKProfile.photo_200.label("vk_photo"))
        .subquery()
    )
    tt_avatar = (
        select(TikTokProfile.account_id.label("tt_account_id"), TikTokProfile.avatar_url.label("tt_photo"))
        .subquery()
    )
    ig_avatar = (
        select(InstagramProfile.account_id.label("ig_account_id"), InstagramProfile.avatar_url.label("ig_photo"))
        .subquery()
    )

    stmt = select(
        SocialAccount,
        metrics_agg.c.views_24h,
        metrics_agg.c.views_7d,
        metrics_agg.c.posts_total,
        metrics_agg.c.last_post_at,
        Phone,
        Email,
        func.max(AccountMetricsDaily.subs).label("subs_total"),
        func.max(AccountMetricsDaily.views).label("views_total"),
        func.max(AccountMetricsDaily.posts).label("videos_total"),
        vk_posts_max.c.vk_last_post_at,
        yt_avatar.c.yt_thumb,
        vk_avatar.c.vk_photo,
        tt_avatar.c.tt_photo,
        ig_avatar.c.ig_photo,
    ).join(
        metrics_agg,
        metrics_agg.c.m_account_id == SocialAccount.id,
        isouter=True,
    ).join(Phone, Phone.id == SocialAccount.phone_id, isouter=True).join(Email, Email.id == SocialAccount.email_id, isouter=True).join(
        AccountMetricsDaily,
        AccountMetricsDaily.account_id == SocialAccount.id,
        isouter=True,
    ).join(
        vk_posts_max,
        vk_posts_max.c.vk_account_id == SocialAccount.id,
        isouter=True,
    ).join(yt_avatar, yt_avatar.c.yt_account_id == SocialAccount.id, isouter=True).join(
        vk_avatar, vk_avatar.c.vk_profile_account_id == SocialAccount.id, isouter=True
    ).join(tt_avatar, tt_avatar.c.tt_account_id == SocialAccount.id, isouter=True).join(
        ig_avatar, ig_avatar.c.ig_account_id == SocialAccount.id, isouter=True
    ).group_by(
        SocialAccount.id,
        metrics_agg.c.views_24h,
        metrics_agg.c.views_7d,
        metrics_agg.c.posts_total,
        metrics_agg.c.last_post_at,
        Phone.id,
        Email.id,
        vk_posts_max.c.vk_last_post_at,
        yt_avatar.c.yt_thumb,
        vk_avatar.c.vk_photo,
        tt_avatar.c.tt_photo,
        ig_avatar.c.ig_photo,
    )

    if platform:
        stmt = stmt.where(SocialAccount.platform == platform)

    if q:
        pattern = f"%{q.lower()}%"
        stmt = stmt.where(
            func.lower(SocialAccount.label).like(pattern)
            | func.lower(SocialAccount.handle).like(pattern)
            | func.lower(SocialAccount.login).like(pattern)
        )

    stmt = stmt.order_by(SocialAccount.created_at.desc())
    result = await session.execute(stmt)
    rows = result.all()
    items: list[SocialAccountRead] = []
    for (
        account,
        views_24h,
        views_7d,
        posts_total,
        last_post_at,
        phone,
        email,
        subs_total,
        views_total,
        videos_total,
        vk_last_post_at,
        yt_thumb,
        vk_photo,
        tt_photo,
        ig_photo,
    ) in rows:
        sync_status = None
        if account.platform == SocialPlatform.youtube and (yt_thumb or account.youtube_channel_id):
            sync_status = "ok"
        if account.platform == SocialPlatform.vk and vk_photo:
            sync_status = "ok"
        if account.platform == SocialPlatform.tiktok and tt_photo:
            sync_status = "ok"
        if account.platform == SocialPlatform.instagram and ig_photo:
            sync_status = "ok"
        items.append(
            SocialAccountRead(
                id=account.id,
                project_id=account.project_id,
                platform=account.platform,
                label=account.label,
                handle=account.handle,
                login=account.login or account.handle,
                url=account.url,
                views_24h=views_24h,
                views_7d=views_7d,
                posts_total=posts_total,
                last_post_at=(vk_last_post_at or last_post_at).isoformat() if (vk_last_post_at or last_post_at) else None,
                phone=PhoneRead.model_validate(phone) if phone else None,
                email=EmailRead.model_validate(email) if email else None,
                youtube_channel_id=account.youtube_channel_id,
                subscribers=subs_total,
                views_total=views_total,
                videos_total=None if account.platform == SocialPlatform.vk else videos_total,
                purchase_price=float(account.purchase_price) if account.purchase_price is not None else None,
                purchase_currency=account.purchase_currency,
                avatar_url=yt_thumb or vk_photo or tt_photo or ig_photo,
                sync_status=account.sync_status or sync_status,
                sync_error=account.sync_error,
                last_synced_at=account.last_synced_at.isoformat() if account.last_synced_at else None,
            )
        )
    return items


@router.post("", response_model=SocialAccountRead, status_code=status.HTTP_201_CREATED)
async def create_account(payload: SocialAccountCreate, session: SessionDep) -> SocialAccountRead:
    if payload.platform == SocialPlatform.vk:
        login = normalize_vk_slug(payload.login or str(payload.url or ""))
    else:
        login = normalize_login(payload.login)
    handle = (payload.handle.strip() if payload.handle else f"@{login}")
    if not handle.startswith("@"):
        handle = f"@{handle}"
    url = build_url(payload.platform, login, str(payload.url) if payload.url else None)

    account = SocialAccount(
        platform=payload.platform,
        label=payload.label.strip(),
        handle=handle,
        login=login,
        url=url,
        account_password=payload.password,
        purchase_price=payload.purchase_price,
        purchase_currency=payload.purchase_currency,
    )
    account.onboarding = AccountOnboarding(status="completed")

    session.add(account)

    sync_status: str | None = None
    sync_error: str | None = None

    try:
        await session.commit()
    except IntegrityError as exc:
        await session.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Account with this login already exists",
        ) from exc

    await session.refresh(account)

    if account.platform == SocialPlatform.youtube and payload.auto_sync:
        try:
            result = await sync_youtube_account(session, account)
            sync_status = result.get("status") or "ok"
            sync_error = result.get("error")
        except HTTPException as exc:
            sync_status = "error"
            sync_error = str(exc.detail)
        except Exception:
            sync_status = "error"
            sync_error = "YouTube sync failed"
    elif account.platform == SocialPlatform.vk and payload.auto_sync:
        try:
            await sync_vk_account(session, account)
            sync_status = "ok"
        except HTTPException as exc:
            sync_status = "error"
            sync_error = str(exc.detail)
        except Exception:
            sync_status = "error"
            sync_error = "VK sync failed"
    elif account.platform == SocialPlatform.tiktok and payload.auto_sync:
        try:
            result = await sync_tiktok_account(session, account)
            sync_status = result.get("status") or "ok"
            sync_error = result.get("error")
        except HTTPException as exc:
            sync_status = "error"
            sync_error = str(exc.detail)
        except Exception:
            sync_status = "error"
            sync_error = "TikTok sync failed"
    elif account.platform == SocialPlatform.instagram and payload.auto_sync:
        try:
            result = await sync_instagram_account(session, account)
            sync_status = result.get("status") or "ok"
            sync_error = result.get("error")
        except HTTPException as exc:
            sync_status = "error"
            sync_error = str(exc.detail)
        except Exception:
            sync_status = "error"
            sync_error = "Instagram sync failed"
    else:
        sync_status = "skipped"

    account_read = await get_account(account.id, session)
    account_read.sync_status = sync_status
    account_read.sync_error = sync_error
    return account_read


@router.get("/{account_id}", response_model=SocialAccountRead)
async def get_account(account_id: int, session: SessionDep) -> SocialAccountRead:
    metrics_agg = (
        select(
            AccountMetricsDaily.account_id.label("m_account_id"),
            func.sum(
                func.coalesce(
                    case(
                        (AccountMetricsDaily.date >= func.current_date() - 1, AccountMetricsDaily.views),
                        else_=0,
                    ),
                )
            ).label("views_24h"),
            func.sum(
                func.coalesce(
                    case(
                        (AccountMetricsDaily.date >= func.current_date() - 7, AccountMetricsDaily.views),
                        else_=0,
                    ),
                )
            ).label("views_7d"),
            func.sum(func.coalesce(AccountMetricsDaily.posts, 0)).label("posts_total"),
            func.max(AccountMetricsDaily.date).label("last_post_at"),
        )
        .where(AccountMetricsDaily.account_id == account_id)
        .group_by(AccountMetricsDaily.account_id)
        .subquery()
    )

    vk_posts_max = (
        select(VKPost.account_id.label("vk_account_id"), func.max(VKPost.published_at).label("vk_last_post_at"))
        .group_by(VKPost.account_id)
        .subquery()
    )
    yt_avatar = (
        select(YouTubeChannel.account_id.label("yt_account_id"), YouTubeChannel.thumbnail_url.label("yt_thumb"))
        .subquery()
    )
    vk_avatar = (
        select(VKProfile.account_id.label("vk_profile_account_id"), VKProfile.photo_200.label("vk_photo"))
        .subquery()
    )
    tt_avatar = (
        select(TikTokProfile.account_id.label("tt_account_id"), TikTokProfile.avatar_url.label("tt_photo"))
        .subquery()
    )
    ig_avatar = (
        select(InstagramProfile.account_id.label("ig_account_id"), InstagramProfile.avatar_url.label("ig_photo"))
        .subquery()
    )

    stmt = (
        select(
            SocialAccount,
            metrics_agg.c.views_24h,
            metrics_agg.c.views_7d,
            metrics_agg.c.posts_total,
            metrics_agg.c.last_post_at,
            Phone,
            Email,
            func.max(AccountMetricsDaily.subs).label("subs_total"),
            func.max(AccountMetricsDaily.views).label("views_total"),
            func.max(AccountMetricsDaily.posts).label("videos_total"),
            vk_posts_max.c.vk_last_post_at,
            yt_avatar.c.yt_thumb,
            vk_avatar.c.vk_photo,
            tt_avatar.c.tt_photo,
            ig_avatar.c.ig_photo,
        )
        .join(metrics_agg, metrics_agg.c.m_account_id == SocialAccount.id, isouter=True)
        .join(Phone, Phone.id == SocialAccount.phone_id, isouter=True)
        .join(Email, Email.id == SocialAccount.email_id, isouter=True)
        .join(AccountMetricsDaily, AccountMetricsDaily.account_id == SocialAccount.id, isouter=True)
        .join(vk_posts_max, vk_posts_max.c.vk_account_id == SocialAccount.id, isouter=True)
        .join(yt_avatar, yt_avatar.c.yt_account_id == SocialAccount.id, isouter=True)
        .join(vk_avatar, vk_avatar.c.vk_profile_account_id == SocialAccount.id, isouter=True)
        .join(tt_avatar, tt_avatar.c.tt_account_id == SocialAccount.id, isouter=True)
        .join(ig_avatar, ig_avatar.c.ig_account_id == SocialAccount.id, isouter=True)
        .where(SocialAccount.id == account_id)
        .group_by(
            SocialAccount.id,
            metrics_agg.c.views_24h,
            metrics_agg.c.views_7d,
            metrics_agg.c.posts_total,
            metrics_agg.c.last_post_at,
            Phone.id,
            Email.id,
            vk_posts_max.c.vk_last_post_at,
            yt_avatar.c.yt_thumb,
            vk_avatar.c.vk_photo,
            tt_avatar.c.tt_photo,
            ig_avatar.c.ig_photo,
        )
    )
    result = await session.execute(stmt)
    row = result.first()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Account not found")
    (
        account,
        views_24h,
        views_7d,
        posts_total,
        last_post_at,
        phone,
        email,
        subs_total,
        views_total,
        videos_total,
        vk_last_post_at,
        yt_thumb,
        vk_photo,
        tt_photo,
        ig_photo,
    ) = row
    return SocialAccountRead(
        id=account.id,
        platform=account.platform,
        label=account.label,
        handle=account.handle,
        login=account.login or account.handle,
        url=account.url,
        views_24h=views_24h,
        views_7d=views_7d,
        posts_total=posts_total,
        last_post_at=(vk_last_post_at or last_post_at).isoformat() if (vk_last_post_at or last_post_at) else None,
        phone=PhoneRead.model_validate(phone) if phone else None,
        email=EmailRead.model_validate(email) if email else None,
        youtube_channel_id=account.youtube_channel_id,
        subscribers=subs_total,
        views_total=views_total,
        videos_total=None if account.platform == SocialPlatform.vk else videos_total,
        purchase_price=float(account.purchase_price) if account.purchase_price is not None else None,
        purchase_currency=account.purchase_currency,
        avatar_url=yt_thumb or vk_photo or tt_photo or ig_photo,
        sync_status=account.sync_status,
        sync_error=account.sync_error,
        last_synced_at=account.last_synced_at.isoformat() if account.last_synced_at else None,
    )


@router.patch("/{account_id}", response_model=SocialAccountRead)
async def update_account(account_id: int, payload: SocialAccountUpdate, session: SessionDep) -> SocialAccountRead:
    account = await session.get(SocialAccount, account_id)
    if not account:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Account not found")

    updates = payload.model_dump(exclude_unset=True)
    if "login" in updates and updates["login"] is not None:
        if account.platform == SocialPlatform.vk:
            updates["login"] = normalize_vk_slug(str(updates["login"]))
        else:
            updates["login"] = normalize_login(str(updates["login"]))
    for field, value in updates.items():
        setattr(account, field, value)

    if account.platform == SocialPlatform.vk:
        account.url = build_url(account.platform, account.login, updates.get("url") if "url" in updates else account.url)
        if account.handle and not account.handle.startswith("@"):
            account.handle = f"@{account.handle}"
        if not account.handle:
            account.handle = f"@{account.login}"
    elif "url" not in updates:
        account.url = build_url(account.platform, account.login, account.url)

    await session.commit()
    await session.refresh(account)
    return await get_account(account_id, session)


@router.get("/export")
async def export_accounts(session: SessionDep):
    result = await session.execute(
        select(SocialAccount, Phone, Email).join(Phone, Phone.id == SocialAccount.phone_id, isouter=True).join(
            Email, Email.id == SocialAccount.email_id, isouter=True
        )
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "accounts"
    headers = [
        "platform",
        "label",
        "handle",
        "url",
        "phone_number",
        "email",
        "email_password",
        "account_password",
        "purchase_source_url",
        "raw_import_blob",
    ]
    ws.append(headers)
    for account, phone, email in result.all():
        ws.append(
            [
                account.platform,
                account.label,
                account.handle,
                account.url,
                phone.phone_number if phone else None,
                email.email if email else None,
                email.email_password if email else None,
                account.account_password,
                account.purchase_source_url,
                account.raw_import_blob,
            ]
        )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=accounts.xlsx"},
    )


@router.post("/import")
async def import_accounts(session: SessionDep, file: UploadFile = File(...)):
    content = await file.read()
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Empty file")
    header = [h.strip() if isinstance(h, str) else h for h in rows[0]]
    idx = {name: i for i, name in enumerate(header)}

    def col(name: str) -> int | None:
        return idx.get(name)

    created = updated = errors = 0
    for row in rows[1:]:
        try:
            handle = row[col("handle")] if col("handle") is not None else None
            if not handle:
                errors += 1
                continue
            handle = str(handle).strip()
            stmt = select(SocialAccount).where(SocialAccount.handle == handle)
            existing = await session.execute(stmt)
            account = existing.scalar_one_or_none()
            phone_number = row[col("phone_number")] if col("phone_number") is not None else None
            email_value = row[col("email")] if col("email") is not None else None
            email_password = row[col("email_password")] if col("email_password") is not None else None
            account_password = row[col("account_password")] if col("account_password") is not None else None
            purchase_source_url = row[col("purchase_source_url")] if col("purchase_source_url") is not None else None
            raw_import_blob = row[col("raw_import_blob")] if col("raw_import_blob") is not None else None
            label = row[col("label")] if col("label") is not None else handle
            platform = row[col("platform")] if col("platform") is not None else "YouTube"
            url = row[col("url")] if col("url") is not None else ""

            phone_obj = None
            email_obj = None
            if phone_number:
                phone_res = await session.execute(select(Phone).where(Phone.phone_number == str(phone_number)))
                phone_obj = phone_res.scalar_one_or_none()
                if not phone_obj:
                    phone_obj = Phone(phone_number=str(phone_number))
                    session.add(phone_obj)
                    await session.flush()
            if email_value:
                email_res = await session.execute(select(Email).where(Email.email == str(email_value)))
                email_obj = email_res.scalar_one_or_none()
                if not email_obj:
                    email_obj = Email(email=str(email_value), email_password=str(email_password) if email_password else None)
                    session.add(email_obj)
                    await session.flush()
                elif email_password:
                    email_obj.email_password = str(email_password)

            if account:
                account.label = str(label)
                account.platform = platform
                account.url = str(url)
                account.account_password = str(account_password) if account_password else None
                account.purchase_source_url = str(purchase_source_url) if purchase_source_url else None
                account.raw_import_blob = str(raw_import_blob) if raw_import_blob else None
                account.phone_id = phone_obj.id if phone_obj else None
                account.email_id = email_obj.id if email_obj else None
                updated += 1
            else:
                account = SocialAccount(
                    platform=platform,
                    label=str(label),
                    handle=handle,
                    url=str(url),
                    account_password=str(account_password) if account_password else None,
                    purchase_source_url=str(purchase_source_url) if purchase_source_url else None,
                    raw_import_blob=str(raw_import_blob) if raw_import_blob else None,
                    phone_id=phone_obj.id if phone_obj else None,
                    email_id=email_obj.id if email_obj else None,
                )
                session.add(account)
                created += 1
            await session.flush()
        except Exception:
            errors += 1
            await session.rollback()
        else:
            await session.commit()
    return {"created": created, "updated": updated, "errors": errors}


@phones_router.get("", response_model=list[PhoneRead])
async def list_phones(session: SessionDep) -> list[Phone]:
    result = await session.execute(select(Phone))
    return result.scalars().all()


@phones_router.post("", response_model=PhoneRead, status_code=status.HTTP_201_CREATED)
async def create_phone(payload: PhoneCreate, session: SessionDep) -> Phone:
    phone = Phone(
        phone_number=payload.phone_number.strip(),
        label=payload.label,
        notes=payload.notes,
    )
    session.add(phone)
    try:
        await session.commit()
    except IntegrityError as exc:
        await session.rollback()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Phone already exists") from exc
    await session.refresh(phone)
    return phone


@emails_router.get("", response_model=list[EmailRead])
async def list_emails(session: SessionDep) -> list[Email]:
    result = await session.execute(select(Email))
    return result.scalars().all()


@emails_router.post("", response_model=EmailRead, status_code=status.HTTP_201_CREATED)
async def create_email(payload: EmailCreate, session: SessionDep) -> Email:
    email = Email(
        email=payload.email.strip(),
        email_password=payload.email_password,
        notes=payload.notes,
    )
    session.add(email)
    try:
        await session.commit()
    except IntegrityError as exc:
        await session.rollback()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Email already exists") from exc
    await session.refresh(email)
    return email
